import openpyxl 
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from datetime import datetime
import os

class Section:
    def __init__(self, qtr, number, title, is_dl, is_async, currics, iname):
        self.number = number
        self.title = title
        self.qtr = qtr
        self.is_dl = is_dl
        self.is_async = is_async
        self.currics = currics
        self.instructor = iname

def is_in_sections(sections, s):
    '''  Is the sectin in a list of sections? '''
    for sec in sections:
        if ((sec.number == s.number) and
            (sec.qtr == s.qtr) and
            (sec.is_dl == s.is_dl) and
            (sec.is_async == s.is_async)):
            return True
    return False

def add_section(sections, qtr, number, title, is_dl, is_async, curric, iname):         
        # Create a section - counts on a sections list at is modified in place
        s = Section(qtr, number, title, is_dl, is_async, [], iname)
        if not is_in_sections(sections, s):
            sections.append(s)

# Open the excel sheet       
fname = 'MAE_MasterCoursePlan_AY25.xlsx'
pdir = '/home/bsb/Downloads'
#planner = '/Users/brianbingham/Downloads/
planner = os.path.join(pdir, fname)

#planner = '//Downloads/MAE_MasterCoursePlan_AY25.xlsx'
#planner='/home/bsb/Downloads/MAE_Master_Course_Plan_AY241.xlsx'
#planner='./MAE_Master_Course_Plan_AY25.xlsx'


wb = openpyxl.load_workbook(planner)
ws = wb['AY25']
#ws = wb.active


# Determine which sections are DL
# Specify the row value that contains "DL" - then read them into a list.
dl = 5
for row in ws.iter_rows(
        min_row = dl, max_row = dl,
        min_col=1, max_col=ws.max_column,
        values_only=True):
    dl_row = row

# Merged dl_row
# Where the cells in the DL row are merged columns,
# repeat the content in each column.
mdl_row = []
for c, col_index in zip(dl_row,range(ws.max_column)):
    val = c
    if c is None:
        # Replace None with blank string
        val = ''
        for crange in ws.merged_cells:
            clo,rlo,chi,rhi = crange.bounds
            top_value = ws.cell(rlo,clo).value
            if (rlo<=dl and dl<=rhi
                and clo<=col_index and col_index<=chi):
                val = top_value
                break
    mdl_row.append(val)
    
# Specify the four rows for the AY we are interested in.
fall_row = 13
summer_row = 16

# Read rows into a list
rows = []
for row in ws.iter_rows(
        min_row = fall_row, max_row = summer_row,
        min_col=1, max_col=ws.max_column,
        values_only=True):
    rows.append(row)

# Keys are instructor names, values are a list of Sections
instructors = dict()

# List of all Sections
section_list = []

qtrs = ['fall', 'winter', 'spring', 'summer']

for row, qtr in zip(rows, qtrs):
    for cell, dl_cell in zip(row, mdl_row):
        if not (cell is None):
            n = cell.find('[')
            m = cell.find(']')
            if ((n > 0) and (m > 0)):                    
                inames = cell[n+1:m]
                # Check for multiple instructors
                iname_list = inames.split('/')
                for iname in iname_list:
                    part = cell[:n-1]
                    ss = part.splitlines()
                    if len(ss) < 1:
                        print("Error with cell: "+cell)
                        break
                    cnum = ss[0]
                    if len(iname_list) > 1:
                        cnum = cnum + '(1/%d)'%len(iname_list)
                    if len(ss) > 1:
                        ctitle = "".join(ss[1:])
                    else:
                        ctitle = ""

                    # DL or not?
                    isdl = False
                    if dl_cell.find('DL') > 0:
                        isdl = True
                    is_async = False
                    if dl_cell.find('Self Paced') > 0:
                        is_async = True
                    
                    # Add a section to the instructors dictionary - if not already there.
                    instructors.setdefault(iname, [])
                    add_section(instructors[iname], qtr, cnum, ctitle,
                                isdl, is_async, [], iname)
                    
                    # Add section to comprehensive list if not already there
                    s = Section(qtr, cnum, ctitle, isdl, is_async, [], iname)
                    if not is_in_sections(section_list, s):
                        section_list.append(s)

# Print to terminal
for k in instructors.keys():
    print(k, end=": ")
    sections = instructors[k]
    for qtr in qtrs:
        print(qtr, end=": ")
        for s in sections:
            if s.qtr == qtr:
                if s.is_dl:
                    print(s.number, end="(DL), ")
                else:
                    print(s.number, end=", ")                    

    print("")
    
                    
# Write teaching schedule to new workbook
tsb = openpyxl.Workbook()
ts = tsb.active

keys = list(instructors.keys())
keys.sort()

r = 1
c = 1
now = datetime.now()

ts.cell(row = 1, column = 1,
        value = 'Draft Teaching Schedule Autogenerated on ' + now.strftime('%d %b %Y %H:%M:%S')   )

r = 2
ts.cell(row = r, column = 1,
        value = 'Name')
ts.cell(row = r, column = 2,
        value = 'Teaching Load, MAE Direct*').alignment = Alignment(wrap_text = True, horizontal = 'center')

for ii in range(1,5):
    ts.merge_cells(start_row = r, start_column = 2*ii+1,
                   end_row = r, end_column = 2*ii + 2)
    ts.cell(row = r, column = 2*ii+1, value = qtrs[ii-1]).alignment = Alignment(horizontal = 'center')

ts.cell(row = r, column = 11, value = 'DL Self Paced').alignment = Alignment(horizontal = 'center')


r = 3
for ii in range(4):
    ts.cell(row = r, column = 2*(ii+1)+1,
            value = 'RES').alignment = Alignment(horizontal = 'center')

    ts.cell(row = r, column = 2*(ii+1)+2,
            value = 'DL').alignment = Alignment(horizontal = 'center')
grey = "E0E0E0"
bg_fill =  PatternFill(start_color=grey, end_color=grey, fill_type = "solid")

r = 4
for k in keys:
    ts.cell(row = r, column = 1,
            value = k).alignment = Alignment(vertical = 'center')
    c = 3
    # Starting row for this instructor
    r0 = r
    maxr = r
    sections = instructors[k]
    for qtr in qtrs:
        r = r0
        for is_dl in [False, True]:
            r = r0
            for s in sections:
                if ((s.qtr == qtr) and (s.is_dl == is_dl) and (not s.is_async)):
                    ts.cell(row = r, column = c,
                            value = s.number).fill = bg_fill
                    if r > maxr:
                        maxr = r
                    r += 1
            c += 1
    # Then right Async to the right
    r = r0
    for s in sections:
        if s.is_async:
            ts.cell(row = r, column = c, value = s.number).fill = bg_fill
            if r > maxr:
                maxr = r
            r += 1
    
    # When instructors have more than one row of classes, merge rows     
    ts.merge_cells(start_row = r0, start_column = 1,
                   end_row = maxr, end_column = 1)
    ts.merge_cells(start_row = r0, start_column = 2,
                   end_row = maxr, end_column = 2)
    
    r = maxr + 1

# Add a note
ts.cell(row = maxr+2, column = 1,
        value = "* Teaching Load, MAE Direct is the number of `segment equivalents` planned in the labor budget.  Typically hybrid (resident+DL concurrenetly) is 1.5 segment equivalents.").alignment = Alignment(wrap_text = False, horizontal = 'left')

# Col width
ts.row_dimensions[2].height = 30

# Set widths
ts.column_dimensions['A'].width = 20
ts.column_dimensions['B'].width = 17

cs = ['C','D','E','F','G','H','I','J','K']
for c in cs:
    ts.column_dimensions[c].width = 17

# Border
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for rr in range(2, r):
    for cc in range(1, 11):
        ts.cell(row=rr, column=cc).border = thin_border

left_border = Border(left=Side(style='medium'), 
                     right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))
right_border = Border(left=Side(style='thin'), 
                     right=Side(style='medium'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))

for rr in range(2, r):
    for cc in [3, 5, 7, 9]:
        ts.cell(row=rr, column=cc).border = left_border
    for cc in [10, 11]:
        ts.cell(row=rr, column=cc).border = right_border
    

# Save
datestamp = datetime.now().strftime("%Y%m%d")
tsb.save('teaching_schedule_%s.xlsx'%datestamp)


# Generate a full list Resident classes
#tlb = openpyxl.Workbook()
#tl = tlb.active

# Sort list by qtr then numbers
qorder = ['fall', 'winter', 'spring', 'summer']
qorder_d = {qtr: index for index, qtr in enumerate(qorder)}
sorted_sections = sorted(section_list, key=lambda section: (qorder_d[section.qtr], section.number))

for s in sorted_sections:
    async_str = 'Async' if s.is_async else ''
    dl_str = 'DL' if s.is_dl else 'RES'
    print(6*"%s, "%(s.qtr, s.number, s.title, dl_str, async_str, s.instructor))


        
