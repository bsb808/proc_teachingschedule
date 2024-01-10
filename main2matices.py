import openpyxl 
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime

def copy_cells_with_formatting(source_file, destination_file, source_sheet_name, destination_sheet_name, source_range, destination_start_cell):
	pass

# Example usage
source_file = './MAE_Master_Course_Plan_AY241.xlsx'
destination_file = 'matrix.xlsx'
source_sheet_name = 'AY24'
destination_sheet_name = 'AY24'
source_range = 'A1:C5'  # Specify the range of cells to copy, adjust as needed
destination_start_cell = (1, 1)  # Specify the starting cell in the destination workbook

#copy_cells_with_formatting(source_file_path, destination_file_path, source_sheet_name, destination_sheet_name, source_range, destination_start_cell)

# Open the source workbook
source_workbook = openpyxl.load_workbook(source_file)
source_sheet = source_workbook[source_sheet_name]

# Open the destination workbook
#destination_workbook = openpyxl.load_workbook(destination_file)
destination_workbook = openpyxl.Workbook()
#destination_sheet = destination_workbook[destination_sheet_name]
destination_sheet = destination_workbook.active

# Define a named style to preserve formatting
cell_style = NamedStyle(name="cell_style")

# Copy cell values and formatting
for row_index, row in enumerate(source_sheet[source_range]):
	for col_index, source_cell in enumerate(row):

		row_i = row=destination_start_cell[0] + row_index
		col_i = destination_start_cell[1] + col_index
		destination_cell = destination_sheet.cell(row=row_i, column=col_i, value = source_cell.value)

		destination_sheet.column_dimensions[get_column_letter(col_i+1)].width = source_sheet.column_dimensions[get_column_letter(col_index+1)].width
		destination_sheet.column_dimensions[get_column_letter(col_i+1)].width = 100

		if source_cell.has_style:
			
		
			# Copy cell formatting
			destination_cell.font = copy(source_cell.font)
			destination_cell.border = copy(source_cell.border)
			destination_cell.fill = copy(source_cell.fill)
			destination_cell.number_format = copy(source_cell.number_format)
			destination_cell.alignment = copy(source_cell.alignment)
			destination_cell.protection = copy(source_cell.protection)
			

			#destination_cell._style = copy(source_cell._style)
			# Apply the named style to the cell
			#destination_cell.style = cell_style

# Save the changes to the destination workbook
destination_workbook.save(destination_file)

# Close the workbooks
source_workbook.close()
destination_workbook.close()

