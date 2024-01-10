import openpyxl 
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string



# Example usage
source_file = './MAE_Master_Course_Plan_AY241.xlsx'
destination_file = 'matrixcut.xlsx'
source_sheet_name = 'AY24'
destination_sheet_name = 'AY24'
source_range = 'A1:C5'  # Specify the range of cells to copy, adjust as needed
destination_start_cell = (1, 1)  # Specify the starting cell in the destination workbook

# Open the source workbook
source_workbook = openpyxl.load_workbook(source_file)
source_sheet = source_workbook[source_sheet_name]


# Cols to keep
c_keep = [1]+list(range(column_index_from_string('J'), 
          	             column_index_from_string('M')+1)) + list(range(column_index_from_string('R'),
          	             column_index_from_string('AL')+1))

c_all = list(range(1, source_sheet.max_column + 1))
c_del = [x for x in c_all if x not in c_keep]

# Rows to keep
r_keep = list(range(1,7)) + list(range(9,18))
r_all = list(range(1, source_sheet.max_row +1))
r_del = [x for x in r_all if x not in r_keep]

for i, c in enumerate(c_del):
	source_sheet.delete_cols(c-i,1)
	
for i, r in enumerate(r_del):
	source_sheet.delete_rows(r-i,1)

source_workbook.save("test3.xlsx")
source_workbook.close()

